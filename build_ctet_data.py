#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Parse the CTET Excel question banks in ./cet into JSON for the ctet_questions table.
(Run `node seed_ctet.mjs` afterwards to load these into PostgreSQL.)

This is deliberately separate from build_data.py / seed.mjs: the UPTET pipeline,
its seed_data/ output and its `questions` table are left completely untouched.

Where build_data.py normalizes aggressively (it drops anything ungradable), this
keeps every Excel row and records *why* a row is unusable, so the raw sheets can
be audited from SQL later.

Outputs:
  seed_data_ctet/manifest.json          -> per-file summary + counts
  seed_data_ctet/rows/<folder>.json     -> every parsed row for one source folder

Run:  python build_ctet_data.py   (or `npm run setup:ctet` to also seed the DB)
"""
import os, re, json, sys
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
CET_DIR = os.path.join(ROOT, "cet")
OUT_DIR = os.path.join(ROOT, "seed_data_ctet")
ROW_DIR = os.path.join(OUT_DIR, "rows")

LETTERS = ["A", "B", "C", "D"]

# The four source folders, and the paper/stream each one represents.
FOLDERS = [
    ("paper 1", "ctet", "paper1", None),
    ("paper2 science", "ctet", "paper2", "science"),
    ("paper2 sst", "ctet", "paper2", "sst"),
    # Not CTET: TGT/PGT/Assistant-Professor Hindi literature PYQs. Loaded under
    # its own exam_id so it is present but trivially separable.
    ("hindi", "hindi-sahitya", None, None),
]

# Sheets that hold commentary rather than questions.
SKIP_SHEETS = {
    "breakdown", "block breakdown", "notes", "answer key", "summary",
    "info", "index", "cover-index",
}

# Byte-identical re-exports of another workbook in the same folder (verified by
# comparing question sets). Kept, but flagged so they can be excluded in one clause.
DUPLICATE_FILES = {
    "CTET_2016_Feb_Paper2_MathsScience_Extracted (1).xlsx",
    "CTET_2023_Aug_Paper2_SST_Extracted (1).xlsx",
    "CTET_2024_15DEC_Paper2_SST_extracted (1).xlsx",
    "Ras_Kavyashastra_PYQ (1).xlsx",
}

# Raw OCR dump: 504 of 606 rows have no usable options. Parsed and stored like
# everything else, but every row lands with is_gradable = false.
KNOWN_BAD = {"Hindi_Sahitya_Adhunik_Kal_OCR_draft.xlsx"}

# One export wrote its data rows one column right of the header from 'Question'
# onward: 'Question' is empty, the question text sits under 'Option_A' and the
# answer letter lands in 'Summary'. The columns stay in header order, so nudging
# just the content indices along recovers every row.
CONTENT_SHIFT = {"Vyakaran_Tadbhav_Tatsam_Sankar.xlsx": 1}


def slug(s):
    s = re.sub(r"[^a-z0-9]+", "-", (s or "").lower()).strip("-")
    return s or "x"


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def combine(a, b):
    """Combine an English/Hindi pair into one display string.

    Uses ' / ' because lib/lang.js splits bilingual text on /\\s+\\/\\s+/.
    """
    a, b = clean(a), clean(b)
    if a and b and a != b:
        return a + " / " + b
    return a or b


def parse_correct(raw, options):
    """Extract the answer letter (A-D) from a messy answer cell.

    Same rules as build_data.py so both banks grade identically.
    """
    s = clean(raw)
    if not s:
        return None
    m = re.match(r"^\(?\s*([ABCDabcd])\s*\)?\s*([:\.\)\-–—]|$)", s)
    if m:
        return m.group(1).upper()
    m = re.search(r"option\s*([ABCDabcd])", s, re.I)
    if m:
        return m.group(1).upper()
    m = re.match(r"^([ABCDabcd])\b", s)
    if m:
        return m.group(1).upper()
    target = re.sub(r"\s+", "", s.lower())
    best, best_letter = 0, None
    for L in LETTERS:
        opt = re.sub(r"\s+", "", clean(options.get(L)).lower())
        if not opt:
            continue
        if opt in target or target in opt:
            if len(opt) > best:
                best, best_letter = len(opt), L
    return best_letter


# ---------------------------------------------------------------- headers ---

# Column names that mark a row as the real header row.
HEADER_HINTS = {
    "question", "question (english)", "q.no", "q.no.", "q_no", "s_no", "row_id",
    "subject", "section", "part", "correct_answer", "correct answer", "answer",
    "option_a", "option (a)", "option a", "option a (english)", "topic",
}


def find_header(ws, maxscan=6):
    """Return (row_number, {normalized header: column index}).

    Several workbooks put a title banner on row 1, so the header row is found by
    scanning rather than assumed.
    """
    best = (None, {}, 0)
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=maxscan, values_only=True), start=1):
        m, score = {}, 0
        for idx, h in enumerate(row):
            h = re.sub(r"\s+", " ", clean(h)).lower()
            if not h:
                continue
            m.setdefault(h, idx)
            if h in HEADER_HINTS:
                score += 1
        if score > best[2]:
            best = (i, m, score)
    return best[0], best[1]


def find(m, *names):
    for nm in names:
        if nm in m:
            return m[nm]
    return None


def cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return clean(row[idx])


def detect_schema(m):
    """Classify a header map into one of the six layouts in the cet/ folder."""
    if "question (english)" in m and "options (english)" in m:
        return "E"   # bilingual, all four options in a single cell
    if "question (english)" in m:
        return "D"   # bilingual, paired columns, one sheet per subject
    if "option (a)" in m or ("section" in m and ("q.no" in m or "q.no." in m)):
        return "C"   # 'solved paper' export
    if "row_id" in m and "version_note" in m:
        return "A"   # standard extraction export
    if "row_id" in m:
        return "F"   # hindi-literature export (Topic / Source_Exam)
    if "s_no" in m or "subject" in m:
        return "B"   # compact export
    return "G"       # anything else; parsed with the generic column aliases


# ------------------------------------------------------------ canonicalise ---

def language_code(raw):
    """Normalize ~15 spellings of the Language column into one token."""
    n = clean(raw).lower()
    if not n or n in ("n/a", "na", "-"):
        return "na"
    has_hi = "hindi" in n or "हिन" in n
    has_en = "english" in n or "eng/" in n or "अंग्रेज" in n
    if "bilingual" in n or (has_hi and has_en):
        return "bilingual"
    if "sanskrit" in n or "संस्कृत" in n:
        return "sanskrit"
    if has_hi:
        return "hindi"
    if has_en:
        return "english"
    return "na"


LANG_SUBJECT = {
    "hindi": "Hindi",
    "english": "English",
    "sanskrit": "Sanskrit",
}


def canon_subject(raw, lang_raw, folder):
    """Map a raw Subject/Section/Part label onto a canonical subject name.

    Two folder-specific rules matter:
      * 'Language-I' / 'Language-II' name no language at all -- the Language
        column has to resolve them.
      * every 'Environmental Studies' label in paper2 sst is really the Social
        Studies section (verified against the question text); Paper 2 has no EVS.
    """
    n = re.sub(r"\s+", " ", clean(raw)).lower()

    if "child" in n or "pedagog" in n or "बाल" in n or re.search(r"\bcdp\b", n):
        return "Child Development & Pedagogy"
    if "sanskrit" in n or "संस्कृत" in n:
        return "Sanskrit"
    if "hindi" in n or "हिन्दी" in n or "हिंदी" in n:
        return "Hindi"
    if "english" in n or "अंग्रेज" in n:
        return "English"
    # 'social' is tested before 'science' so that the very common label
    # 'Social Studies / Social Science' does not fall through to Maths & Science.
    if "social" in n or "सामाजिक" in n:
        return "Social Studies"
    if "science" in n or "विज्ञान" in n:
        return "Mathematics & Science"
    if "math" in n or "गणित" in n:
        return "Mathematics & Science" if folder != "paper 1" else "Mathematics"
    if "environment" in n or "evs" in n or "पर्यावरण" in n:
        return "Social Studies" if folder == "paper2 sst" else "Environmental Studies"
    if re.search(r"language\s*-?\s*(i{1,2}|1|2)\b", n):
        mapped = LANG_SUBJECT.get(language_code(lang_raw))
        if mapped:
            return mapped
    return clean(raw)


# Workbooks with no Subject/Section/Part column at all: the subject has to come
# from the printed question number. Bands below were read off the question text
# at each boundary -- note this paper runs CDP/Maths/EVS *before* the languages,
# which is not the order the CTET blueprint prints, so this cannot be generalized.
QNO_BANDS = {
    "CTET_2023_Paper1_Questions.xlsx": [
        (1, 30, "Child Development & Pedagogy"),
        (31, 60, "Mathematics"),
        (61, 90, "Environmental Studies"),
        (91, 120, None),   # Language I  -- language read from the question text
        (121, 150, None),  # Language II -- language read from the question text
    ],
}

# Language-paper rows in those workbooks are prefixed '[English]', '[हिन्दी …]',
# '[संस्कृत …]'. Blocks are contiguous, so an untagged row inherits the last tag.
LANG_TAG = re.compile(r"^\s*\[([^\]]{0,40})\]")


def tag_language(question):
    m = LANG_TAG.match(question or "")
    if not m:
        return None
    tag = m.group(1).lower()
    if "संस्कृत" in tag or "sanskrit" in tag:
        return "Sanskrit"
    if "हिन्दी" in tag or "हिंदी" in tag or "hindi" in tag:
        return "Hindi"
    if "english" in tag:
        return "English"
    return None


def band_subject(filename, s_no, question, carried):
    """Subject for a workbook that has no subject column. Returns (name, carried)."""
    bands = QNO_BANDS.get(filename)
    if not bands or not re.fullmatch(r"\d+", clean(s_no)):
        return None, carried
    n = int(s_no)
    for lo, hi, name in bands:
        if lo <= n <= hi:
            if name:
                return name, carried
            found = tag_language(question) or carried
            return found, found
    return None, carried


MONTHS = ["jan", "feb", "march", "mar", "apr", "may", "jun", "july", "jul",
          "aug", "sep", "oct", "nov", "dec"]


def parse_exam(filename):
    """Pull (year, session) out of a filename like CTET_2024_Dec_Paper1..."""
    year = None
    m = re.search(r"(19|20)\d{2}", filename)
    if m:
        year = int(m.group(0))
    session = None
    low = filename.lower()
    for mon in MONTHS:
        if re.search(r"[_\s\-]" + mon, low):
            session = mon.capitalize()
            break
    return year, session


# ----------------------------------------------------------------- parsing ---

OPT_SPLIT = re.compile(r"\(\s*([abcd])\s*\)\s*", re.I)


def split_options_blob(text):
    """Schema E keeps all four options in one cell: '(a) x\\n(b) y\\n(c) z'."""
    out = {L: "" for L in LETTERS}
    s = clean(text)
    if not s:
        return out
    parts = OPT_SPLIT.split(s)
    # parts = [pre, letter, body, letter, body, ...]
    for i in range(1, len(parts) - 1, 2):
        L = parts[i].upper()
        if L in out:
            out[L] = parts[i + 1].strip().strip("\n")
    return out


def parse_sheet(ws, sheet_name, ctx):
    """Yield one dict per Excel row in a question sheet."""
    hrow, m = find_header(ws)
    if not m:
        return []
    schema = detect_schema(m)

    q_i = find(m, "question", "प्रश्न")
    q_en_i = find(m, "question (english)")
    q_hi_i = find(m, "प्रश्न (हिन्दी)", "question (hindi)")
    if q_i is None and q_en_i is None:
        return []

    # answer: prefer an explicit letter column, keep the prose one separately
    if "key" in m:
        ans_i = m["key"]
        text_i = find(m, "correct option text", "answer")
    else:
        ans_i = find(m, "correct_answer", "correct answer", "answer / उत्तर", "answer")
        text_i = find(m, "correct option text")

    exp_i = find(m, "explanation", "summary", "explanation_summary",
                 "explanation / व्याख्या (हिन्दी)")
    exp_en_i = find(m, "explanation (english)")
    exp_hi_i = find(m, "व्याख्या (हिन्दी)", "explanation (hindi)")

    opt_i = {L: find(m, f"option_{L.lower()}", f"option ({L.lower()})",
                     f"option {L.lower()}") for L in LETTERS}
    opt_en_i = {L: find(m, f"option {L.lower()} (english)") for L in LETTERS}
    opt_hi_i = {L: find(m, f"विकल्प {L.lower()} (हिन्दी)") for L in LETTERS}
    blob_en_i = find(m, "options (english)")
    blob_hi_i = find(m, "विकल्प (हिन्दी)")

    # Meta columns stay put; only the content columns are offset.
    shift = CONTENT_SHIFT.get(ctx["file"], 0)
    if shift:
        sh = lambda i: None if i is None else i + shift
        q_i, q_en_i, q_hi_i = sh(q_i), sh(q_en_i), sh(q_hi_i)
        ans_i, text_i = sh(ans_i), sh(text_i)
        exp_i, exp_en_i, exp_hi_i = sh(exp_i), sh(exp_en_i), sh(exp_hi_i)
        opt_i = {L: sh(v) for L, v in opt_i.items()}
        opt_en_i = {L: sh(v) for L, v in opt_en_i.items()}
        opt_hi_i = {L: sh(v) for L, v in opt_hi_i.items()}
        blob_en_i, blob_hi_i = sh(blob_en_i), sh(blob_hi_i)

    meta = {
        "row_id": find(m, "row_id"),
        "s_no": find(m, "s_no", "q.no", "q.no.", "q_no", "q.no. / प्र.सं."),
        "part": find(m, "part"),
        "section_label": find(m, "section", "section / अनुभाग"),
        "subject_raw": find(m, "subject"),
        "language_raw": find(m, "language"),
        "version_note": find(m, "version_note"),
        "topic": find(m, "topic"),
        "sub_theme": find(m, "sub_theme"),
        "source_exams": find(m, "source_exams", "source_exam"),
        "scan_note": find(m, "scan note"),
    }

    out = []
    carried = None  # last language seen, for workbooks resolved by question number
    for excel_row, row in enumerate(ws.iter_rows(min_row=hrow + 1, values_only=True),
                                    start=hrow + 1):
        q_en = cell(row, q_en_i)
        q_hi = cell(row, q_hi_i)
        question = cell(row, q_i) or combine(q_en, q_hi)
        if not question:
            continue

        # options: one column per letter, a bilingual pair, or a single blob
        if blob_en_i is not None or blob_hi_i is not None:
            en_opts = split_options_blob(cell(row, blob_en_i))
            hi_opts = split_options_blob(cell(row, blob_hi_i))
            opts = {L: combine(en_opts[L], hi_opts[L]) for L in LETTERS}
        elif any(opt_en_i[L] is not None for L in LETTERS):
            opts = {L: combine(cell(row, opt_en_i[L]), cell(row, opt_hi_i[L]))
                    for L in LETTERS}
        else:
            opts = {L: cell(row, opt_i[L]) for L in LETTERS}

        correct_raw = cell(row, ans_i)
        correct = parse_correct(correct_raw, opts)

        # why a row cannot be scored; mirrors build_data.py's skip conditions
        skip = None
        if sum(1 for v in opts.values() if v) < 2:
            skip = "fewer than 2 options"
        elif correct is None:
            skip = "no parsable answer"
        elif not opts.get(correct):
            skip = "answer points at an empty option"

        explanation = cell(row, exp_i) or combine(cell(row, exp_en_i), cell(row, exp_hi_i))
        raw = {k: (cell(row, i) or None) for k, i in meta.items()}

        subject_raw = raw["subject_raw"] or raw["part"] or raw["section_label"] or raw["topic"]
        if ctx["schema_subject_from_sheet"]:
            subject_raw = sheet_name
        if subject_raw:
            subject_name = (subject_raw if ctx["folder"] == "hindi"
                            else canon_subject(subject_raw, raw["language_raw"], ctx["folder"]))
        else:
            subject_name, carried = band_subject(ctx["file"], raw["s_no"], question, carried)

        out.append({
            **raw,
            "source_folder": ctx["folder"],
            "source_file": ctx["file"],
            "source_sheet": sheet_name,
            "source_row": excel_row,
            "schema_variant": schema,
            "exam_id": ctx["exam_id"],
            "paper": ctx["paper"],
            "stream": ctx["stream"],
            "exam_year": ctx["year"],
            "exam_session": ctx["session"],
            "subject_raw": subject_raw or None,
            "subject_id": slug(subject_name) if subject_name else None,
            "subject_name": subject_name or None,
            "language_code": language_code(raw["language_raw"]),
            "question": question,
            "question_en": q_en or None,
            "question_hi": q_hi or None,
            "option_a": opts["A"] or None,
            "option_b": opts["B"] or None,
            "option_c": opts["C"] or None,
            "option_d": opts["D"] or None,
            "correct": correct,
            "correct_raw": correct_raw or None,
            "correct_text": cell(row, text_i) or None,
            "explanation": explanation or None,
            "explanation_en": cell(row, exp_en_i) or None,
            "explanation_hi": cell(row, exp_hi_i) or None,
            "options_en_raw": cell(row, blob_en_i) or None,
            "options_hi_raw": cell(row, blob_hi_i) or None,
            "is_gradable": skip is None,
            "skip_reason": skip,
            "is_duplicate_file": ctx["file"] in DUPLICATE_FILES,
            "qkey": re.sub(r"\s+", "", question.lower())[:200],
        })
    return out


def process_file(path, folder, exam_id, paper, stream):
    name = os.path.basename(path)
    year, session = parse_exam(name)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        _, m = find_header(ws)
        # Schema D splits one paper across per-subject sheets, so the sheet name
        # is the only place the subject is recorded.
        ctx = {
            "folder": folder, "file": name, "exam_id": exam_id, "paper": paper,
            "stream": stream, "year": year, "session": session,
            "schema_subject_from_sheet": detect_schema(m) == "D",
        }
        try:
            rows.extend(parse_sheet(ws, sheet_name, ctx))
        except Exception as e:
            print(f"    !! {name} / {sheet_name}: {e}", file=sys.stderr)
    wb.close()
    return rows


def main():
    os.makedirs(ROW_DIR, exist_ok=True)
    manifest = {"folders": [], "totals": {}}
    grand = {"rows": 0, "gradable": 0, "files": 0}

    for folder, exam_id, paper, stream in FOLDERS:
        d = os.path.join(CET_DIR, folder)
        if not os.path.isdir(d):
            print(f"skip missing folder: {folder}")
            continue
        files = sorted(f for f in os.listdir(d) if f.lower().endswith(".xlsx")
                       and not f.startswith("~$"))
        all_rows, file_stats = [], []
        for f in files:
            rows = process_file(os.path.join(d, f), folder, exam_id, paper, stream)
            all_rows.extend(rows)
            ok = sum(1 for r in rows if r["is_gradable"])
            file_stats.append({
                "file": f, "rows": len(rows), "gradable": ok,
                "schemas": sorted({r["schema_variant"] for r in rows}),
                "duplicate_of_another_file": f in DUPLICATE_FILES,
                "known_bad_ocr": f in KNOWN_BAD,
            })

        out_name = slug(folder) + ".json"
        with open(os.path.join(ROW_DIR, out_name), "w", encoding="utf-8") as fh:
            json.dump(all_rows, fh, ensure_ascii=False)

        ok = sum(1 for r in all_rows if r["is_gradable"])
        manifest["folders"].append({
            "folder": folder, "exam_id": exam_id, "paper": paper, "stream": stream,
            "file": out_name, "files": file_stats,
            "rows": len(all_rows), "gradable": ok,
            "unique_questions": len({r["qkey"] for r in all_rows}),
        })
        grand["rows"] += len(all_rows)
        grand["gradable"] += ok
        grand["files"] += len(files)
        print(f"{folder:<16} files={len(files):<3} rows={len(all_rows):<6} gradable={ok}")

    manifest["totals"] = grand
    with open(os.path.join(OUT_DIR, "manifest.json"), "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, ensure_ascii=False, indent=2)

    print(f"\nTotal: {grand['files']} files, {grand['rows']} rows, "
          f"{grand['gradable']} gradable -> {OUT_DIR}")


if __name__ == "__main__":
    main()
