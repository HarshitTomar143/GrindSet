#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Convert the UPTET Excel question banks into JSON consumed by the Next.js app.

Outputs:
  public/data/manifest.json                  -> nav tree (sections/groups/subjects + counts)
  public/data/questions/<section>__<subject>.json -> full question list per subject

Run:  python build_data.py
"""
import os, re, json, math, sys
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(ROOT, "public", "data")
Q_DIR = os.path.join(OUT_DIR, "questions")
MOCK_SIZE = 30

LETTERS = ["A", "B", "C", "D"]


def slug(s):
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return s or "x"


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def classify(sheet_name, is_science=False):
    """Map a sheet name to a canonical subject name."""
    n = sheet_name.lower()
    if "child" in n or "बाल" in n:
        return "Child Development & Pedagogy"
    if "sanskrit" in n or "संस्कृत" in n:
        return "Sanskrit"
    if "math" in n or "गणित" in n:
        return "Mathematics & Science" if is_science else "Mathematics"
    if "hindi" in n or "हिन्दी" in n or "हिंदी" in n:
        return "Hindi"
    if "english" in n or "अंग्रेज" in n:
        return "English"
    if "environment" in n or "पर्यावरण" in n:
        return "Environmental Studies"
    if "social" in n or "सामाजिक" in n:
        return "Social Studies"
    return sheet_name.strip()


def header_map(ws):
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    m = {}
    for i, h in enumerate(row):
        if h is None:
            continue
        m[str(h).strip()] = i
    return m


def find(m, *names):
    """Return column index for the first header name that exists."""
    for nm in names:
        if nm in m:
            return m[nm]
    return None


def cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return clean(row[idx])


def combine(a, b):
    """Combine a Hindi/English pair into one display string."""
    a, b = clean(a), clean(b)
    if a and b and a != b:
        return a + "  /  " + b
    return a or b


def parse_correct(raw, options):
    """Extract the answer letter (A-D) from a messy 'Correct Answer' cell."""
    s = clean(raw)
    if not s:
        return None
    # leading letter, optionally bracketed, then a separator or end
    m = re.match(r"^\(?\s*([ABCDabcd])\s*\)?\s*([:\.\)\-–—]|$)", s)
    if m:
        return m.group(1).upper()
    # "Option X ..."
    m = re.search(r"option\s*([ABCDabcd])", s, re.I)
    if m:
        return m.group(1).upper()
    # bare leading letter followed by a word boundary
    m = re.match(r"^([ABCDabcd])\b", s)
    if m:
        return m.group(1).upper()
    # fuzzy: match the answer text against the option texts
    target = re.sub(r"\s+", "", s.lower())
    best, best_letter = 0, None
    for L in LETTERS:
        opt = re.sub(r"\s+", "", clean(options.get(L)).lower())
        if not opt:
            continue
        if opt and (opt in target or target in opt):
            score = len(opt)
            if score > best:
                best, best_letter = score, L
    return best_letter


def extract_sheet(ws, is_science):
    """Yield normalized question dicts from a worksheet."""
    m = header_map(ws)
    q_hi = find(m, "Question (Hindi)", "Question")
    q_en = find(m, "Question (English)")
    correct_i = find(m, "Correct Answer")
    exp_hi = find(m, "Explanation (Hindi)", "Explanation")
    exp_en = find(m, "Explanation (English)")
    opt_hi = {L: find(m, f"Option {L} (Hindi)", f"Option {L}") for L in LETTERS}
    opt_en = {L: find(m, f"Option {L} (English)") for L in LETTERS}

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        question = combine(cell(row, q_hi), cell(row, q_en))
        if not question:
            continue
        opts = {}
        opts_present = 0
        for L in LETTERS:
            txt = combine(cell(row, opt_hi[L]), cell(row, opt_en[L]))
            opts[L] = txt
            if txt:
                opts_present += 1
        if opts_present < 2:
            continue
        correct = parse_correct(cell(row, correct_i), opts)
        if correct is None or not opts.get(correct):
            continue  # skip ungradable questions to keep scoring honest
        explanation = combine(cell(row, exp_hi), cell(row, exp_en))
        out.append({
            "question": question,
            "options": opts,
            "correct": correct,
            "explanation": explanation,
        })
    return out


def process_file(path, is_science=False):
    """Return {canonical_subject: [questions...]} merged across sheets."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    subjects = {}
    for sh in wb.sheetnames:
        ws = wb[sh]
        subject = classify(sh, is_science)
        subjects.setdefault(subject, []).extend(extract_sheet(ws, is_science))
    wb.close()
    # de-duplicate identical questions within a subject
    for subj, qs in subjects.items():
        seen, uniq = set(), []
        for q in qs:
            key = re.sub(r"\s+", "", q["question"].lower())
            if key in seen:
                continue
            seen.add(key)
            uniq.append(q)
        subjects[subj] = uniq
    return subjects


def build_section(section_id, section_name, groups_spec):
    """groups_spec: list of (group_id, group_name, subjects_dict)"""
    groups = []
    for gid, gname, subjects in groups_spec:
        subj_list = []
        for subject_name in sorted(subjects.keys()):
            qs = subjects[subject_name]
            if not qs:
                continue
            sid = slug(subject_name)
            fname = f"{section_id}__{gid}__{sid}.json"
            with open(os.path.join(Q_DIR, fname), "w", encoding="utf-8") as f:
                json.dump(qs, f, ensure_ascii=False)
            subj_list.append({
                "id": sid,
                "name": subject_name,
                "total": len(qs),
                "mocks": max(1, math.ceil(len(qs) / MOCK_SIZE)),
                "file": fname,
            })
        groups.append({"id": gid, "name": gname, "subjects": subj_list})
    return {"id": section_id, "name": section_name, "groups": groups}


def main():
    os.makedirs(Q_DIR, exist_ok=True)

    p1 = process_file(os.path.join(ROOT, "paper1_section_wise.xlsx"))
    sci = process_file(os.path.join(ROOT, "paper2_science.xlsx"), is_science=True)
    sst = process_file(os.path.join(ROOT, "paper2_sst.xlsx"))

    manifest = {
        "mockSize": MOCK_SIZE,
        "sections": [
            build_section("paper1", "Paper 1", [
                ("main", "Subjects", p1),
            ]),
            build_section("paper2", "Paper 2", [
                ("science", "Science Stream (Maths & Science)", sci),
                ("sst", "Social Studies Stream", sst),
            ]),
        ],
    }
    with open(os.path.join(OUT_DIR, "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    # summary
    print("Built data:")
    for sec in manifest["sections"]:
        print(f"\n# {sec['name']}")
        for g in sec["groups"]:
            print(f"  [{g['name']}]")
            for s in g["subjects"]:
                print(f"    - {s['name']}: {s['total']} questions -> {s['mocks']} mock paper(s)")


if __name__ == "__main__":
    main()
